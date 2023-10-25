import time

import os
from openpyxl import Workbook,load_workbook

# @Time    : {2023/5/24} {21:33}
# @Author  : backperker
# @File    : D:\python\PythonProject\SummarizeProject\ReadXlsx.py
# @Description : #作用筛选一个表的重复数据，并选择创建先得表把数据保存
#@path xlsx 路径
#@sheetname  表名
def readXlsx(path,sheetName):
    if not os.path.exists(path):
        print("文件不存在")
        return
    wb = load_workbook(path)
    sheets = wb.sheetnames
    sheetIndex=-1

    for itemIndex in range(len(sheets)):
        print(f"{sheetName}=={sheets[itemIndex]}")
        if sheetName==sheets[itemIndex]:
            sheetIndex=itemIndex
            break

    if sheetIndex==-1:
        print(f"xlsx 表格没有此{sheetName}")
        return

    sh = wb[sheets[sheetIndex]]
    rows = sh.rows
    columns = sh.columns
    print(f"{rows},{columns}")
    listsNew=[]
    repeatlists=[]
    for row in rows:
        # print(row)
        row_val=[col.value for col in row].__getitem__(0)
        if row_val  in listsNew:
            repeatlists.append(row_val)
        else:
            listsNew.append(row_val)
        # print(row_val)
    print(repeatlists)
    saveRepeatData(path,repeatlists)
    # for column in columns:
    #     print(column)
    #     column_val=[col.value for col in column]
    #     print(column_val)
#@path 保存的路径
#@list 保存重复数据数组
def saveRepeatData(path,list):
    if  len(list)==0:return
    wb = load_workbook(path)
    timedata=time.strftime("%H-%M-%S",time.localtime())
    wb.create_sheet(f"重复数据{timedata}")
    sh = wb[f"重复数据{timedata}"]
    index=1
    for item in list:
        sh.cell(row=index,column=1).value=item
        index+=1
    wb.save(path)
    wb.close()
    print(f"保存重复的数据{timedata}")


def init():
    xlsxPath = "D:\python\记录.xlsx"
    sheetName = "Sheet1"
    readXlsx(xlsxPath,sheetName)

if __name__=="__main__":
    init()