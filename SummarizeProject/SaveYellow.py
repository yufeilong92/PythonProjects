import time
# @Time    : {2023/5/24} {21:38}
# @Author  : backperker
# @File    : D:\python\PythonProject\SummarizeProject\SaveYellow.py
# @Description : 用于保存固定网址数据，并形成xslx
import requests
import os
from bs4 import BeautifulSoup


from openpyxl import Workbook, load_workbook

#@xlsxpath xlsx 地址
#@sheetaname 表名
#@pageindex 页面索引
#@url 地址
#@baseheadr  基础base

def saveYellow(xlsxPath, sheetName, pageindex, url, baseHear):
    # ===================创建文件开始===================
    row_num = 1
    column_num = 1
    # 判断文件是否存在
    if not os.path.exists(xlsxPath):
        print("文件已经存在"+f'{xlsxPath}')
        wb = Workbook()
        sh = wb.create_sheet(sheetName)
    else:
        # 存在
        wb = load_workbook(xlsxPath)
        sheets = wb.sheetnames
        indexPostion = -1
        print(sheets)
        for index in range(len(sheets)):
            if sheets[index] == sheetName:
                indexPostion = index
                break
        if indexPostion == -1:
            print("不存在，创建 "+f'{sheetName}')
            wb.create_sheet(sheetName)
            sh = wb[sheetName]
        else:
            print(f'{sheetName}{indexPostion}'+"已存在该表格，追加")
            sh = wb[sheets[indexPostion]]
            # 行---
            row_num = sh.max_row + 1
            # ||| 列
            # column_num = sh.max_cplumn

    # =================创建文件结束================
    #      if row_num == 1:
    #          print(f'当前序号{row_num}')
    #          sh.cell(row=1, column=column_num).value = "序号"
    #          sh.cell(row=1, column=column_num + 1).value = "名称"
    #          sh.cell(row=1, column=column_num + 2).value = "地址"
    #          sh.cell(row=row_num + 1, column=column_num).value = f'{row_num + 1}'
    #          sh.cell(row=row_num + 1, column=column_num + 1).value = f"{saveWebTitle}"
    #          sh.cell(row=row_num + 1, column=column_num + 2).value = f"{saveWebUrl}"
    #      else:
    #          print(f'当前序号{row_num}')
    #          sh.cell(row=row_num, column=column_num).value = f'{row_num + 1}'
    #          sh.cell(row=row_num, column=column_num + 1).value = f"{saveWebTitle}"
    #          sh.cell(row=row_num, column=column_num + 2).value = f"{saveWebUrl}"
    # ===================请求网络开始==============

    # 时间戳
    # sheetTime=time.strftime("%H-%M-%S",time.localtime())
    requestPageNum = 1
    baseUrl = url
    requestPage = 1
    # sheetRoWPage=1
    if row_num==1:
        sh.cell(row=1, column=column_num).value = "序号"
        sh.cell(row=1, column=column_num + 1).value = "名称"
        sh.cell(row=1, column=column_num + 2).value = "地址"
        row_num+=1
    print(f"行数开头：{row_num}")
    while  requestPageNum<=pageindex:
        url=f'{baseUrl}&page={requestPageNum}'
        print(url)
        req = requests.get(url)
        req.encoding = "utf-8"
        rep = BeautifulSoup(req.text, "html.parser")
        find_all = rep.findAll("h3", class_="")
        if not len(find_all):
            print("未找到数据")
            break
        for reqitem in find_all:
            saveWebUrl=baseHear+reqitem.find('a')['href']
            saveWebTitle=reqitem.getText()
            sh.cell(row=row_num, column=column_num).value = f'{row_num}'
            sh.cell(row=row_num, column=column_num + 1).value = f"{saveWebTitle}"
            sh.cell(row=row_num, column=column_num + 2).value = f"{saveWebUrl}"
            row_num +=1
        requestPageNum +=1
    else:
        print("已经请求完成，Complate")
    # sh.cell(row=row_num,column=column_num).value="值1"

    # 保存
    wb.save(xlsxPath)
    wb.close()

    contion = input("是否继续，按enter，继续，否则输入任何内容，结束")
    if not contion:
        init()
    else:
        print("结束,over")

def init():
    path = input("请输入保存文件内容，默认D:\python\数据.xlsx\n")
    if not path:
        #保存地址
        xlsxPath = "D:\python\数据.xlsx"
    else:
        xlsxPath=str(path)
    webHear = input("请输入网页头，默认：https://z2212k.xyz/pw/\n")
    if not webHear:
        #网页头
        baseHear = "https://z2212k.xyz/pw/"
    else:
        baseHear=str(webHear)
    weburl = input("请输入网址地址，默认：https://z2212k.xyz/pw/thread1022.php?fid=198\n")
    if not weburl:
        #网址
        url = "https://z2212k.xyz/pw/thread1022.php?fid=198"
    else:
        url=str(weburl)
    sheet = input("请输入sheel名称，默认sheet\n")
    if not sheet:
        #表名
        sheetName = "sheet"
    else:
        sheetName=str(sheet)
    num=input("请输入循环多少页,默认0,请输入数字\n")
    if not num:
        #循环多少次
        pageindex =0
    else:
        pageindex=int(num)
    saveYellow(xlsxPath, sheetName, pageindex, url, baseHear)


if __name__ == "__main__":
    init()