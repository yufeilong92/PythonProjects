# @Time    : {2023/5/25} {21:39}
# @Author  : backperker
# @File    : D:\python\PythonProject\SummarizeProject\ReadAndCopyXlsx.py
# @Description : 读取现有xlsx ,并拷贝重复数据
import os.path
import time

from DataVo import DataVo
from Utils import Utils
from openpyxl import Workbook, load_workbook


class ReadAndCopyXlsx:
    def initInput(self):
        u = Utils()
        vo = DataVo()
        print("请选择XLSX文件，其他文件报错")
        time.sleep(1)
        readPath = u.selectFiledocumentDialog()
        print(f"你选择{readPath}")
        readSheel = u.inputUtils(vo.PLEASE_READ_SHEET)
        self.readAndCopy(readPath, readSheel)

    def readAndCopy(self, readPath, readSheet, u):
        if not os.path.exists(readPath):
            print(f"没有该文件{readPath}")
            return

        wb = load_workbook(readPath)
        sheetnames = wb.sheetnames
        isExices = -1
        for index in range(len(sheetnames)):
            if sheetnames[index] == readSheet:
                isExices = index
                break
        if isExices == -1:
            print(f"{readPath}文件中没有该{readSheet}表")
            return
        sh = wb[sheetnames[isExices]]
        rows = sh.rows
        print("查询结果中....")
        for row in rows:
            rows_ = [col.value for col in row]
            print(rows_)
        self.selectItem(readPath,rows, wb, sh, u)

    def selectItem(self,readPath, rows, wb, sh, u):
        isSelect = u.inputUtils("1 选择拷贝某列数据 \n"
                                "2 查找某列列表重复数据 \n"
                                "3 保存某列的数据重复数据 \n"
                                "4 保存某列去重后的数据 \n")
        while not isSelect.isdigit():
            isSelect = u.inputUtils("1 选择拷贝某列数据 \n"
                                    "2 查找某列列表重复数据 \n"
                                    "3 保存某列的数据重复数据 \n"
                                    "4 保存某列去重后的数据 \n")
        if isSelect == 1:
            print()
            self.selectOneItme(readPath,rows, wb, sh, u)
        elif isSelect == 2:
            print()
        elif isSelect == 3:
            print()
        elif isSelect == 4:
            print()
        else:
            self.selectItem(rows, wb, sh, u)

    def selectOneItme(self,readPath, rows, wb, sh, u):
        # 选择拷贝某列数据
        columns = sh.columns
        selectColumn = u.inputUtils('请输入选择的第几行数据')

        while not selectColumn.isdigit() or int(selectColumn) > len(columns) + 1:
            selectColumn = u.inputUtils('请输入选择的第几行数据')

        selectRows = set()
        for row in rows:
            value = [col.value for col in row].__getitem__(selectColumn - 1)
            selectRows.add(value)
        print(f'你选择的{selectColumn}列的数据是{selectRows}')

        isSave = input("是否选择保存数据，enter 保存，输入任何数据退出")

        if isSave:
            print("不保存，退出")
            return
        else:
            print("保存数据中")
            if len(selectRows) == 0:
                print("要保存的数据是空的")
                return
            else:
                createTime = time.strftime("%Y-%M-%d %H:%M:%S", time.localtime())
                wb.create_sheet(createTime)
                sh = wb[createTime]
                for itemIndex in range(len(selectRows)):
                    sh.cell(itemIndex, 1).value = selectRows[itemIndex]
                wb.save(readPath)
                wb.close()