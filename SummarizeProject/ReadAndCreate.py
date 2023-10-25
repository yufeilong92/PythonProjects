#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/5/26 21:12
# @Author  : backpacker
# @File    : ReadAndCreate.py
# @Description :读取已有的xlsx 并把数据追加到另个xlsx
import os.path

from Dialog import DialogSHow
from Utils import Utils
from tkinter import *
from openpyxl import load_workbook, Workbook
from tkinter import messagebox

class ReadAndCreate:
    def appendDialogXlsx(self,root, xlsx_Path, var, wb, sh, rows, u, data):
        row_number = sh.max_row + 1
        root.destroy()
        # 要保存的列数
        index = var.get()
        for item in data:
            row_number += 1
            sh.cell(row=row_number, column=int(index)).value = f"{item}"
        wb.save(xlsx_Path)
        wb.close()
        askyesno = messagebox.askyesno("询问", message="是否关闭程序？")
        if askyesno :
            exit()



    def appendSelectSheet(self,xlsx_Path, wb: Workbook, sheetName: str, sh, u: Utils, data):
        rows = sh.rows
        columns = sh.max_column
        # 弹窗选择
        root = Tk()
        root.title("请选择操作流程")
        var = IntVar(master=root)
        var.set(0)
        root.geometry("650x450")
        index = 0
        for prompt in rows:
            for prov in prompt:
                name = f"第{index}列，值={prov.value}"
                print(name)
                radiobutton = Radiobutton(root, text=name, font="18",indicatoron=False,bg="light gray", selectcolor="light green", variable=var, value=index)
                radiobutton.pack(fill=X, ipady=5)
                if index < columns:
                    index += 1
            break
        btn = Button(root, text="确认", font="18",
                     command=lambda: self.appendDialogXlsx(root, xlsx_Path, var, wb, sh, rows, u, data))
        btn.pack(side=TOP, ipady=1)
        root.mainloop()


    def selectFile(self,entry: Entry, labal: Label, utils: Utils):
        strip = entry.get().title().strip()
        if not strip:
            messagebox.showwarning(title="顺序警告", message="请先输入表名")
        else:
            create_path = utils.selectFileDialogWithTitle("请选择目录")
            complate_path = create_path + f"/{strip}.xlsx"
            if os.path.exists(complate_path):
                messagebox.showerror(title="错误提示", message=f"{complate_path}已经存在，请选择目录")
            elif not create_path:
                messagebox.showerror(title="错误提示", message=f"{complate_path}未选择，请选择目录")
            else:
                labal["text"]=f"{create_path}"
            #     s = input("是否重新操作，按enter,继续")
            #     if not s:
            #         createNewXlsx(utils, data)
            #     else:
            #         return
            # else:
            #     createXlsxNew(complate_path, data)


    def CreateXlsx(self,root, entry: Entry,labal_Path:Label,  data):

        strip = entry.get().title().strip()
        create_path = labal_Path["text"]
        if not strip:
            messagebox.showwarning(title="温馨提示", message="请先输入表名")
        elif create_path=="请选择地址":
            messagebox.showwarning(title="温馨提示", message="请选择保存目录")
        else:
            complate_path = create_path + f"/{strip}.xlsx"
            if os.path.exists(complate_path):
                messagebox.showerror(title="错误提示", message=f"{complate_path}已经存在，请重新输入")
                # s = input("是否重新操作，按enter,继续")
                # if not s:
                #     createNewXlsx(utils, data)
                # else:
                #     return
            else:
                root.destroy()
                self.createXlsxNew(complate_path,strip, data)


    def createNewXlsx(self,u, data):
        root = Tk()
        root.geometry("640x540")
        root.title("创建xlsx")
        root.config(bg="SeaShell")
        labal = Label(root, text="表名", font="18")
        labal.grid(row=0, column=1, columnspan=1, sticky=E + W)

        edit = Entry(root, font="18")
        edit.grid(row=0, column=2, columnspan=1, sticky=E + W)
        labal1 = Label(root, text=".xlsx", font="18")
        labal1.grid(row=0, column=3, columnspan=1, sticky=E + W)

        labal_Path = Label(root, text="请选择地址", font="18")
        labal_Path.grid(row=1, column=1, columnspan=1, sticky=E + W)
        btn = Button(root, text="选择地址", font="18", command=lambda: self.selectFile(edit, labal_Path, u))
        btn.grid(row=1, column=2, columnspan=1, sticky=E + W)
        btnsure = Button(root, text="创建", font="18", command=lambda: self.CreateXlsx(root, edit,labal_Path,data))
        btnsure.grid(row=2, column=1, columnspan=2, sticky=E + W)
        root.mainloop()

        # save_sheetName = u.inputStr("请输入要创建的表名")
        # while not save_sheetName.__getitem__(1):
        #     save_sheetName = u.inputStr("请输入要创建的表名")
        # save_path = u.selectFileDialogWithTitle("选择要保存的路径")
        # complate_path = save_path + f"/{save_sheetName.__getitem__(1)}.xlsx"
        # if os.path.exists(complate_path):
        #     messagebox.showinfo(title="错误提示", message=f"{complate_path}已经存在，请重新输入")
        #     s = input("是否重新操作，按enter,继续")
        #     if not s:
        #         createNewXlsx(u, data)
        #     else:
        #         return

        # createXlsxNew(complate_path, data)


    def createXlsxNew(self,complate_path,save_sheetName, data):
        wb = Workbook()
        sh = wb.create_sheet(save_sheetName.__getitem__(0))
        row_number = 0
        for item in data:
            row_number += 1
            sh.cell(row_number, 1).value = f"{item}"
        wb.save(complate_path)
        print(f"{complate_path}保存成功")
        messagebox.showinfo(title="温馨提示",message="创建并保存成功")
        wb.close()

        if messagebox.askyesno("询问","是否结束程序？"):
            exit()
        # sss = input("按Enter 退出 \n")
        # if not sss:
        #     print("程序结束")


    def appendXlsx(self,u, data):
        print("create")
        # 文件路径
        xlsx_Path = u.selectFiledocumentDialogWithTitle("请选择要追加的xlsx文件")
        wb = load_workbook(xlsx_Path)
        sheetnames = wb.sheetnames
        sheet_name = ""
        if len(sheetnames) == 1:
            sheet_name = sheetnames[0]
        else:
            print(f"选择的文件里面的任意表\n{sheetnames}\n")
            sheet_name = u.inputStr("请输入选择的表名")
            # 是否包含
            contain = u.isContain(sheet_name, sheetnames)
            while not contain:
                print("输入的表名,当前文件不存在，请重新输入")
                sheet_name = u.inputStr("请输入选择的表名")
                contain = u.isContain(sheet_name, sheetnames)
        sh = wb[sheet_name]
        row_number = sh.max_row + 1
        column = sh.max_column
        if column <= 1:  # 就一列直接追u加
            for item in data:
                row_number += 1
                sh.cell(row_number, 1).value = f"{item}"

            wb.save(xlsx_Path)
            wb.close()
            print(f"保存成功{xlsx_Path}=={sheet_name}")
            messagebox.showinfo(title="温馨提示", message="追加并保存成功")

            if messagebox.askyesno("询问","是否结束程序？"):
                return
            # sss = input("按Enter 退出 \n")
            # if not sss:
            #     print("程序结束")
            # return
        else:  ##存在多列
            self.appendSelectSheet(xlsx_Path, wb, sheet_name, sh, u, data)


    def dialogSelectOrCreate(self,root, var, u, data):
        root.destroy()
        index = var.get()
        if index == 1:
            self.appendXlsx(u, data)
        elif index == 2:
            self.createNewXlsx(u, data)


    def selectOrCreate(self,u, data: []):
        # 选择追加的xlsx，或者创建xlsx
        root = Tk()
        root.title("请选择操作流程")
        var = IntVar(master=root)
        var.set(1)
        root.config(bg="SeaShell")
        # root.geometry("650x450")
        radiobutton = Radiobutton(root, text="选择已有的xlsx,追加数据", font="18", bg="light gray", selectcolor="light green", variable=var,
                                  value=1,indicatoron=False)
        radiobutton.grid(row=1,columnspan=3, sticky=W+E,pady=4)
        createbutton = Radiobutton(root, text="创建新的xlsx,并保存", font="18",bg="light gray", selectcolor="light green", variable=var,
                                   value=2,indicatoron=False)
        createbutton.grid(row=2, columnspan=3,sticky=W+E,pady=4)
        btn = Button(root, text="确认", font="18", command=lambda: self.dialogSelectOrCreate(root, var, u, data))
        btn.grid(row=3,columnspan=3, sticky=W+E,pady=8)
        root.mainloop()

        # 选择表格，
        # 追加数据


    def readData(self,u: Utils, rows, selectCoulumn):
        readData = []
        for row in rows:
            row_val = [col.value for col in row].__getitem__(int(selectCoulumn))
            readData.append(row_val)
        print(readData)
        self.selectOrCreate(u, readData)


    def selectColumsSure(self,root: Tk, var, wb: Workbook, raw, u: Utils):
        print(f"{var.get()}")
        root.destroy()
        selectCoulumn = var.get()
        wb.close()
        self.readData(u, raw, selectCoulumn)


    def SelectSheet(self,wb: Workbook, sheetName: str, u: Utils):
        sh = wb[sheetName]
        rows = sh.rows
        max_row = sh.max_row
        # 没有数据
        if max_row < 1: return
        columns = sh.max_column
        if columns == 1:  # 就一列数据，无需弹窗
            self.readData(wb, sheetName, u, 0)
            return

        # 弹窗选择
        root = Tk()
        root.title("请选择要保存的列")
        invar = IntVar(master=root)
        root.config(bg="SeaShell")
        # root.geometry("650x450")
        index = 0
        # , bg = "light gray", selectcolor = "light green"
        for prompt in rows:
            for prov in prompt:
                name = f"第{index}列，值={prov.value}"
                print(name+f"\n===={index}")
                Radiobutton(root, text=name,  font="18",bg="light gray", selectcolor="light green",variable=invar,indicatoron=False,value=index).grid(row=index,sticky=W+E,pady=4)
                if index < columns:
                    index += 1

            break
        btn = Button(root, text="确认", font="18", command=lambda: self.selectColumsSure(root, invar, wb, rows, u))
        btn.grid(row=index+2, sticky=W+E,pady=8)
        root.mainloop()


    def readFileAndData(self):
        # 文件绝对路径
        u = Utils()
        filePath = u.selectFiledocumentDialog()
        is_open = u.check_excel_isOpen(filePath)
        isNext = True
        if not filePath:
            messagebox.showwarning("温馨提示","未选择文件")
            return
        while is_open & isNext:
            print(f"{filePath}处于打开状态,请关闭继续执行程序")
            inputNumber = u.inputInt("请输入1 继续，\n其他结束")
            isNext = inputNumber != 1
            if isNext:
                is_open = u.check_excel_isOpen(filePath)
            else:
                exit()
        wb = load_workbook(filePath)
        sheetnames = wb.sheetnames
        sheet_name = ""
        if len(sheetnames) == 1:
            sheet_name = sheetnames[0]
        else:
            print(f"选择的文件里面的任意表\n{sheetnames}\n")
            sheet_name = u.inputStr("请输入选择的表名")
            # 是否包含
            contain = u.isContain(sheet_name, sheetnames)
            while not contain:
                print("输入的表名,当前文件不存在，请重新输入")
                sheet_name = u.inputStr("请输入选择的表名")
                contain = u.isContain(sheet_name, sheetnames)
        self.SelectSheet(wb, sheet_name, u)


    # if __name__ == '__main__':
    #     readFileAndData()
