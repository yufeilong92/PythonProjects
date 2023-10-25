#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/5/24 22:19
# @Author  : 我的名字
# @File    : Test.py
# @Description : $
# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import time
import os
import openpyxl
from selenium.webdriver.common.by import By


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


def getContent(url):
    req = requests.get(url)
    req.encoding = "utf-8"
    r = BeautifulSoup(req.text, "html.parser")
    resut = r.findAll("h3")
    # print(resut)
    for item in resut:
        print(item)


# Press the green button in the gutter to run the script.

def getTitleAndUrl(save, url):
       postion = 1
       baseItemHear = "https://z2212k.xyz/pw/"

       req = requests.get(url)
       req.encoding = "utf-8"
       rep = BeautifulSoup(req.text, "html.parser")
    # prettify = rep.prettify()
       if os.access(save, os.W_OK):
           print("文件已经存在")
       else:
           open(save, 'a+')

       find_all = rep.findAll("h3", class_="")
       print(find_all)
       for item in find_all:
           print(f"==={postion}========")
           saveWebUrl = baseItemHear + item.find('a')['href']
           saveName = item.get_text()
           savexlsx(save,postion,f"{postion}",saveName,saveWebUrl)
        # print(baseItemHear + item.find('a')['href'])
        # print(item.get_text())
        # with open(save, "a+", encoding='utf-8') as f:
        #     f.write(saveName)
        #     f.write("       ")
        #     f.write(saveWebUrl)
        #     f.write("\n")
           postion += 1
           print("=====over======")
    # print(find_all)



def SaveYellowWeb(save: str, url: str):
    baseUrl = url
    postion = 1
    while postion < 44:
        url = f"{baseUrl}&page={postion}"
        getTitleAndUrl(save, url)
        # req=requests.get(url)
        # req.encoding="utf-8"
        # rep=BeautifulSoup(req.text,"html.parser")
        # print(rep.get("a"))
        # values=f"{postion}"
        # chrom_driver = webdriver.Chrome()
        # time.sleep(1)
        # chrom_driver.get(url)
        # chrom_driver.find_element(By.PARTIAL_LINK_TEXT,value=values).click()
        postion += 1
        print(url)
        time.sleep(10)
    else:
        print("faild")


def savexlsx(save, rows, value1, value2, value3):
    if os.access(save, os.W_OK):
            print("文件已经存在")
    else:
            open(save, "a+")

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("yellow")
    print(f"savexlsx=={rows}=={value1}=={value2}=={value3}")
    ws.cell(row=rows, column=1).value = "22"
    ws.cell(row=rows, column=2).value = "22"
    ws.cell(row=rows, column=3).value = "33"
    # 5.保存表格
    wb.save(save)
    time.sleep(1)


def saveXles(save,url,pagename,pagemax,pageindex):
    baseUrl = url
    postionPage = 1
    postionItem=1
    baseItemHear = "https://z2212k.xyz/pw/"
    if os.access(save, os.W_OK):
        print("文件已经存在")
    else:
        open(save, "a+")

    wb = openpyxl.Workbook()
    wb.create_chartsheet()
    ws = wb.create_sheet(pagename,pageindex)
    while postionPage < pagemax:
        url = f"{baseUrl}&page={postionPage}"
        print(url)
        # 开始请求数据
        print("开始请求数据")
        req = requests.get(url)
        req.encoding = "utf-8"
        rep = BeautifulSoup(req.text, "html.parser")
        if os.access(save, os.W_OK):
            print("文件已经存在")
        else:
            open(save, 'a+')

        find_all = rep.findAll("h3", class_="")
        print(find_all)
        for item in find_all:
            print(f"==={postionItem}========")
            saveWebUrl = baseItemHear + item.find('a')['href']
            saveName = item.get_text()
            # 保存
            print("开始保存保存")

            print(f"savexlsx=={postionItem}=={postionItem}=={saveName}=={saveWebUrl}")
            ws.cell(row=postionItem, column=1).value = f'{postionItem}'
            ws.cell(row=postionItem, column=2).value = f'{saveName}'
            ws.cell(row=postionItem, column=3).value = f'{saveWebUrl}'
            postionItem += 1
            print("=====over======")
        time.sleep(1)
        postionPage += 1

    else:
        print("faild")
    # 5.保存表格
    wb.save(save)
    wb.close()

if __name__ == '__main__':
    # print_hi('PyCharm')
    # url = "https://z2212k.xyz/pw/thread1022.php?fid=219&page=1"
    # url = "https://z2212k.xyz/pw/thread1022.php?fid=219"
    url = "https://z2212k.xyz/pw/thread1022.php?fid=5"
    # save="D:\python\save.txt"
    save = "D:\python\save.xlsx"
    pagename="yelww"
    pagemax=2
    pageindex=1
    saveXles(save,url,pagename,pagemax,pageindex)
    # savexlsx(save)

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
