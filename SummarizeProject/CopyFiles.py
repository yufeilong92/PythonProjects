#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/5/28 14:28
# @Author  : backpacker
# @File    : CopyFiles.py
# @Description : 拷贝选择当前的目前所有数据
import os

from tkinter import *
from tkinter import messagebox

from Utils import Utils

from  openpyxl import Workbook

class CopyFiles:
    xx = "xxxxxx"

    def selectFile(self):
        utils = Utils()
        copy_path = utils.selectFileDialogWithTitle("选择要拷贝目录")
        data = os.listdir(copy_path)

    def selectDialog(self):
        utils = Utils()
        root = Tk()
        root.title("选择拷贝目录")
        # root.geometry("640x540")
        root.config(bg="light gray")

        Label(root, text="拷贝目录：", font=14).grid(row=1, pady=4,sticky=W+E+S+N)

        label = Label(root,  font=14)
        label.grid(row=1, column=1, pady=4,sticky=W+E+S+N)

        btn = Button(root, text="选择目录", font=14,  command=lambda: self.selectSureDialog(label, utils))
        btn.grid(row=1, column=2, pady=4,sticky=W+E)

        Label(root, text="拷贝地址：", font=14).grid(row=2, pady=4,sticky=W+E+S+N)
        savelabel = Label(root,  font=14)
        savelabel.grid(row=2, column=1, pady=4,sticky=W+E+S+N)

        savebtn = Button(root, text="选择目录",font=14,  command=lambda: self.saveSureDialog(savelabel, utils))
        savebtn.grid(row=2, column=2, pady=4,sticky=W+E)

        Label(root, text="拷贝名称：", font=14).grid(row=3, pady=4,sticky=W+E+S+N)

        edit = Entry(root, font=14)
        edit.grid(row=3, column=1,  pady=4,sticky=W+E+S+N)


        var = IntVar()
        var.set(1)
        Label(text="导出数据类型", padx="50",font=14, justify=CENTER,width=40).grid(row=6,  pady=4,column=0, columnspan=3,sticky=W+E)
        Radiobutton(root, text="TXT", variable=var,background="gray",selectcolor="green", value=1, indicatoron=False,justify=CENTER).grid(row=7, column=0, columnspan=3,sticky=W+E)
        Radiobutton(root, text="XLSX", variable=var, background="gray",selectcolor="green",value=2, indicatoron=False, justify=CENTER).grid(row=8, column=0, columnspan=3,sticky=W+E)

        createBtn = Button(root, text="创建", font="12",
                           command=lambda: self.createBtnData(var, savelabel, label, edit))
        createBtn.grid(row=12, column=0, pady=18,columnspan=3,sticky=W+E)

        mainloop()

    def createBtnData(self, var: IntVar, savelabel: Label, label: Label, edit: Entry):
        copypath = label["text"]
        print(f"copypath{copypath}")
        savepath = savelabel["text"]
        print(f"savepath{savepath}")
        title = edit.get().title()
        selectType = var.get()
        print(f"type{selectType}")
        if copypath == "" or copypath is None:
            messagebox.showinfo("温馨提示", "请选项拷贝目录")
            return
        if savepath == "" or savepath is None:
            messagebox.showinfo("温馨提示", "请选择保存目录")
            return
        if title == "":
            messagebox.showinfo("温馨提示", "请输入文件名")
            return
        suffix=".txt"
        if selectType == 1:
            suffix = ".txt"
        else:
            suffix=".xlsx"
        complate = savepath +"/"+ title+suffix
        if  os.path.exists(complate):
            messagebox.showinfo("温馨提示",f"{complate}本地存在，请输入文件名")
            return
        data = os.listdir(copypath)

        if selectType == 1:
            txt = open(complate, "w")
            for itme in data:
                txt.write(itme+"\n")
            txt.close()
            messagebox.showinfo("温馨提示","保存成功")
            os.startfile(savepath)
        else:
            wb=Workbook()
            index=1
            sh = wb.create_sheet(title)
            for item in data:
                sh.cell(row=index,column=1).value=f"{item}"
                index+=1
            wb.save(complate)
            wb.close()
            messagebox.showinfo("温馨提示","保存成功")
            os.startfile(savepath)


    def selectSureDialog(self, lab: Label, util: Utils):
        selectPath = util.selectFileDialogWithTitle("请选择要拷贝的目录")
        lab["text"] = f"{selectPath}"

    def saveSureDialog(self, savelabel, utils):
        savePath = utils.selectFileDialogWithTitle("请选择保存的目录")
        savelabel["text"] = f"{savePath}"
