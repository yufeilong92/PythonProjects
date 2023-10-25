import os.path
import time

import requests
from bs4 import BeautifulSoup

from openpyxl import Workbook, load_workbook

import DataVo
import Utils
from selenium import webdriver
from Utils import Utils
from DataVo import DataVo
from selenium.webdriver.common.by import By


class JAV_Web:
    def createXlsx(self, savePath, saveSheetName, weburl, maxPageNum, baseHear):
        row_init = 1
        column_init = 1
        currentNum = 1
        # ================文件保存=========================
        if not os.path.exists(savePath):
            print(f"{savePath}文件不存在，创建")
            wb = Workbook()
            sh = wb.create_sheet(saveSheetName)
        else:
            print(f"{savePath}文件存在")
            wb = load_workbook(savePath)
            sheetnames = wb.sheetnames
            isExistIndex = -1
            # isExist = list(filter(lambda x: x == saveSheetName, sheetnames))
            for existIndex in range(len(sheetnames)):
                if saveSheetName == sheetnames[existIndex]:
                    isExistIndex = existIndex
                    break

            if isExistIndex == -1:  # 没有存在
                print(f'{saveSheetName}：表不存在，创建')
                wb.create_sheet(saveSheetName)
                sh = wb[saveSheetName]
            else:
                print(f"{saveSheetName}：表已经存在，追加")
                sh = wb[sheetnames[isExistIndex]]
                row_init = sh.max_row + 1
                # ----
        # =================启动浏览器========================
        web_driver = webdriver.Firefox()
        isException = Utils.isExceptionGET(web_driver, weburl)

        while not isException:
            isException = Utils.isExceptionGET(web_driver, weburl)
            if not isException:
                time.sleep(10)

        isContinue = Utils.inputUtils(DataVo.PLEASE_INPUT_CONTIUNE)

        if isContinue:
            print("返回了")
            return
        web_driver.refresh()
        isRefresh = Utils.isExceptionRefresh(web_driver)
        while isRefresh:
            isRefresh = Utils.isExceptionRefresh(web_driver)
        print("获取网页信息数据")
        print(f"网页标题{web_driver.title}")

        try:
            while currentNum <= maxPageNum:
                currentNum += 1
                source = web_driver.page_source
                rep = BeautifulSoup(source, "html.praser")
                find_all = rep.findAll('div', class_='video-elem')

                for item in find_all:
                    web_time = item.findNext('small').text
                    # print(f"分钟=={timea.strip()}")
                    web_title = item.findNext('a', class_='title text-sub-title mt-2 mb-3').text
                    # print(f"=标题 ={title_a.strip()}")
                    web_url = item.findNext('a')['href']
                    row_init += 1
                    sh.cell(row_init, 1).value = row_init
                    sh.cell(row_init, 2).value = f"{web_time.strip()}"
                    sh.cell(row_init, 3).value = f"{web_title.strip()}"
                    sh.cell(row_init, 4).value = f"{baseHear}{web_url.strip()}"

                print(f"当前行数{row_init}，当前页面{currentNum}/{maxPageNum}")

                try:
                    find_element = web_driver.find_element(By.LINK_TEXT, '下一页')  # 可用
                    if find_element.is_enabled() == True:
                        find_element.click()
                        print("能点击")
                    else:
                        print("不能点击")
                except Exception as e:
                    print(f"已经是最后一页{e}")
                    break
                else:
                    print("网页执行成功，over")
        except:
            print("网页执行成功，over")
        wb.save(savePath)
        wb.close()
        print("保存成功")
        # =========================================

    def JAV_init(self):
        u=Utils()
        vo=DataVo()

        saveFileName = u.inputUtils(vo.PLEASE_INPUT_SAVE_NAME)
        print("选择要保存的文件夹")
        time.sleep(1)
        path = u.selectFileDialog()
        savePath=f"{path}/{saveFileName}"
        saveSheelName = u.inputUtils(vo.PLEASE_INPUT_SHEET_NAME)
        maxPageNum = u.inputUtils(vo.PLEASE_INPUT_MAX_PAGE)
        baseHear = "https://jav169.top/"
        weburl = u.inputUtils(vo.PLEASE_INPUT_WEBURL)
        self.createXlsx(savePath, saveSheelName, weburl, maxPageNum, baseHear)
