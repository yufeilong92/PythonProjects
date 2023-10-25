import os
import time

from bs4 import BeautifulSoup
import requests
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By

from openpyxl import Workbook, load_workbook


def isinput(str):
    # url = input("请输入网址地址：\n")
    s = input(f"{str}")
    if s:
        return (s, True)
    else:
        return (s, False)


def init(isfirst, opt):
    # url = input("请输入网址地址：\n")
    # if not  url:
    savePath = "D:\python\jav.xlsx"
    # issheetname=input("请输入表名,默认sheet 、\n")
    sheelist = ("", False)
    while not sheelist.__getitem__(1):
        sheelist = isinput("请输入表名 、\n")
        if sheelist.__getitem__(1):
            sheetName = str(sheelist.__getitem__(0))

    # if issheetname:
    #     sheetName=str(issheetname)
    # else:
    #     sheetName = "sheet"

    # isUrl=input("请输入网址，\n")
    weblist = ("", False)
    while not weblist.__getitem__(1):
        weblist = isinput("请输入网址，\n")
        if weblist.__getitem__(1):
            weburl = str(weblist.__getitem__(0))
    # if isUrl:
    #     weburl=str(isUrl)
    # else:
    #     weburl = "https://jav169.top/?0KPHnqeXlKrEmIiupJ91G9v8SSDj"
    # weburl="https://hanime1.me/search?genre=%E8%A3%8F%E7%95%AA&page=2"
    # weburl="https://www.baidu.com/"
    # else:
    #     weburl=url
    baseHear = "https://jav168.top"

    # ismAX=input("请输入最大页码 \n")
    webPageNumlist = ("", False)
    while not webPageNumlist.__getitem__(1):
        webPageNumlist = isinput("请输入最大页码 \n")
        if webPageNumlist.__getitem__(1):
            maxIndex = int(webPageNumlist.__getitem__(0))

    # if ismAX:
    #     maxIndex = int(ismAX)
    # else:
    #     maxIndex=1
    print("执行程序")
    saveXlsx(weburl, savePath, sheetName, baseHear, maxIndex, isfirst, opt)


def saveXlsx(url, savePath, sheetName, baseHear, maxIndex, isfirst, opt):
    row_num = 1
    currentNum = 1
    if not os.path.exists(savePath):
        print(f"{savePath}文件不存在，创建")
        wb = Workbook()
        sh = wb.create_sheet(sheetName)
    else:
        wb = load_workbook(savePath)
        sheetnames = wb.sheetnames
        # 是否要添加新的sheet,默认添加
        sheetIndex = -1
        for indexA in range(len(sheetnames)):
            if sheetnames[indexA] == sheetName:
                sheetIndex = indexA
                break
        if sheetIndex == -1:
            print(f"不存在，创建{sheetName}")
            wb.create_sheet(sheetName)
            sh = wb[sheetName]
        else:
            print(f"已经存在，追加{sheetName}")
            sh = wb[sheetnames[sheetIndex]]
            # |||
            row_num = sh.max_row + 1
    # =================================
    if isfirst:
        option = Options()
        option.set_capability('pageLoadStrategy', 'eager')
        print("开始启动网络")
        opt = webdriver.Firefox(options=option)
    try:
        opt.get(url)
    except Exception as e:
        print("异常继续 继续")

    connect = input("判断是否继续\n")
    if connect:
        print("返回了")
        return

    print("刷新")

    execption = isExecption(opt)

    while execption:
        execption = isExecption(opt)
    # opt.refresh()
    print("获取标题")
    print(opt.title)
    # source = opt.page_source
    # print("获取源码打印中")
    # isconnect = True
    try:
        while currentNum <= maxIndex:
            currentNum += 1
            # print(opt.current_url)
            # =========================
            source = opt.page_source
            rep = BeautifulSoup(source, "html.parser")
            find_all = rep.findAll('div', class_='video-elem')
            print("======webData========")
            for item in find_all:
                # print("======start========")

                timea = item.findNext('small').text
                # print(f"分钟=={timea.strip()}")

                title_a = item.findNext('a', class_='title text-sub-title mt-2 mb-3').text
                # print(f"=标题 ={title_a.strip()}")

                urla = item.findNext('a')['href']
                # print(f"连接=={urla.strip()}")

                itemtitle = item.text
                # print(f"===itemtitle==={itemtitle.strip()}========")
                # print("======over========")
                row_num += 1
                sh.cell(row_num, 1).value = row_num
                sh.cell(row_num, 2).value = f"{title_a.strip()}"
                sh.cell(row_num, 3).value = f"{timea.strip()}"
                sh.cell(row_num, 4).value = f"{baseHear}{urla.strip()}"
            # =========================

            print(f"当前行数{row_num}，当前页面{currentNum}/{maxIndex}")
            # saa = input("判断是否下一页,按Enter 进入下一页\n")
            # if saa:
            #     print("不进行下一页")
            #     isconnect = False
            # else:
            #     print("下一页")
            #     isconnect = True
            #     print("执行点击按钮")
            try:
                find_element = opt.find_element(By.LINK_TEXT, '下一页')  # 可用
                if find_element.is_enabled() == True:
                    find_element.click()
                    print("能点击")
                else:
                    print("不能点击")
            except Exception as e:
                print(f"已经是最后一页{e}")
                break
        else:
            print("已经执行完毕，over")
    except Exception as e:
        print(f"抓取异常保存{e}")

    wb.save(savePath)
    wb.close()
    opt.back()
    print(f"保存成功{savePath}=={sheetName}")
    sss = input("是否继续,按Enter 继续 \n")
    if sss:
        print("程序结束")
    else:
        init(False, opt)


def isExecption(opt):
    try:
        opt.refresh()
        print("执行刷新")

        return False
    except:
        print("刷新异常")
        return True


if __name__ == "__main__":
    init(True, None)
